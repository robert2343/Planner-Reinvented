import openpyxl
import datetime
import DailyTask

def wdts(num):
    if num == 0:
        return "Monday"
    elif num == 1:
        return "Tuesday"
    elif num == 2:
        return "Wednesday"
    elif num == 3:
        return "Thursday"
    elif num == 4:
        return "Friday"
    elif num == 5:
        return "Saturday"
    elif num == 6:
        return "Sunday"

def stwd(weekday):
    if weekday == "Monday":
        return 0
    elif weekday == "Tuesday":
        return 1
    elif weekday == "Wednesday":
        return 2
    elif weekday == "Thursday":
        return 3
    elif weekday == "Friday":
        return 4
    elif weekday == "Saturday":
        return 5
    elif weekday == "Sunday":
        return 6

def initFile(hwfilename, assignmentSheetName, projectSheetName, dailyTaskSheetName, dailyTaskDataSheetName):
    wb = openpyxl.Workbook()
    wb.active.title = assignmentSheetName
    wb.create_sheet().title = projectSheetName
    wb.create_sheet().title = dailyTaskSheetName
    wb.create_sheet().title = dailyTaskDataSheetName
    wb.active['A1'] = "Title"
    wb.active['B1'] = "Category"
    wb.active['C1'] = "Deadline"
    wb.active['D1'] = "Location"
    wb.active['E1'] = "Notes"
    wb.active = wb[projectSheetName]
    wb.active['A1'] = "Title"
    wb.active['B1'] = "Deadline"
    wb.active['C1'] = "Notes"
    wb.active = wb[dailyTaskSheetName]
    wb.active['A1'] = "Title"
    wb.active['B1'] = "Pattern"
    wb.active['C1'] = "Notes"
    wb.active = wb[dailyTaskSheetName]
    wb.active['A1'] = "Title"
    wb.active['B1'] = "Begin Data"
    wb.save(filename=hwfilename)

def printHelp():
    print("-a: add assignment")
    print("-f: remove assignment")
    print("-p: add project")
    print("-i: remove project")
    print("-g: add daily task")
    print("-w: mark off daily task as successful")
    print("-d: remove daily task")
    print("-h: print this help menu")
    print("-v: print your to do list, including daily tasks that you don't need to do today.")
    print("No command line arguments: print your to-do list")
    print("Don't use this unless your computer's clock is correct.")
    print("  On Linux/BSD/MacOS, see the date command.")
    print("  On Windows, look at the \"Change Date and Time\" settings in \"System Settings\".")
    print("To learn more: https://gitlab.com/123456robert/planner-reinvented")

def month2Int(month):
    if month == "January":
        return 1
    elif month == "February":
        return 2
    elif month == "March":
        return 3
    elif month == "April":
        return 4
    elif month == "May":
        return 5
    elif month == "June":
        return 6
    elif month == "July":
        return 7
    elif month == "August":
        return 8
    elif month == "September":
        return 9
    elif month == "October":
        return 10
    elif month == "November":
        return 11
    elif month == "December":
        return 12
    else:
        return -1

def sortDateLambda(date):
    if date == None:
        return datetime.datetime(9999, 12, 31, 23, 59, 59, 999999)
    else:
        return date

def findIndexbyFirstCell(file, cellVal, sheetName):
    wb = openpyxl.load_workbook(filename=file)
    sheet = wb[sheetName]
    iterRows = sheet.iter_rows()
    listRows = list(iterRows)
    for i in range(1, len(listRows)):
        if listRows[i][0].value == cellVal:
            return i

def deleteItemByName(file, sheetName, delUp):
    wb = openpyxl.load_workbook(filename=file)
    sheet = wb[sheetName]
    name = input("Enter the name of the assignment to remove: ")
    idx = findIndexbyFirstCell(file, name, sheetName)
    sheet.delete_rows(idx + 1 - delUp, 1 + delUp)
    wb.save(file)
    return name

def deleteItemByNameNoInp(file, sheetName, name, delUp):
    wb = openpyxl.load_workbook(filename=file)
    sheet = wb[sheetName]
    idx = findIndexbyFirstCell(file, name, sheetName)
    sheet.delete_rows(idx + 1 - delUp, 1 + delUp)
    wb.save(file)
    return idx

#decompose and numToLetter taken from https://codereview.stackexchange.com/questions/182733/base-26-letters-and-base-10-using-recursion
def decompose(num):
    while num:
        num, remainder = divmod(num - 1, 26)
        yield remainder

def numToLetter(num):
    return ''.join(chr(ord('A') + part)for part in decompose(num))[::-1]

def writeCell(x, y, val, file, sheetName):
    wb = openpyxl.load_workbook(filename=file)
    sheet = wb[sheetName]
    sheet[numToLetter(x) + str(y)].value = val
    wb.save(file)

def updateDaily(val, name, file, sheetName, dataSheetName, refDate, suppressOutput):
    idx = findIndexbyFirstCell(file, name, sheetName)
    dt = DailyTask.DailyTask.readInFromFile(idx + 1, file, sheetName, dataSheetName)
    if len(dt.boolArr) >= len(dt.datesArr):
        print("There is a serious problem. If you didn't manually modify the Excel file, report a bug.")
    else:
        if dt.datesArr[-1].replace(hour=0, minute=0, second=0, microsecond=0) == refDate.replace(hour=0, minute=0, second=0, microsecond=0):
            placeAtX = len(dt.boolArr) + 2
            placeAtY = findIndexbyFirstCell(file, name, dataSheetName) + 1
            writeCell(placeAtX, placeAtY, val, file, dataSheetName)
            writeCell(placeAtX + 1, placeAtY - 1, dt.calculateNextDate(refDate + datetime.timedelta(days=1)), file, dataSheetName)
        else:
            if not suppressOutput:
                print("This cannot be checked off today")