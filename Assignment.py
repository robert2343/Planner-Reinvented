import datetime
import helper
import openpyxl

class Assignment:
    def __init__(self, _name, _category, _deadline, _location, _notes):
        self.name = str(_name)
        self.category = str(_category)
        self.deadline = _deadline
        self.location = str(_location)
        self.notes = str(_notes)
        self.none = _name == None

    def __str__(self):
        result = self.name + ", " + self.category
        if datetime.datetime.now() > self.deadline - datetime.timedelta(days=7):
            result += ", " + helper.wdts(self.deadline.weekday()) + " " + str(self.deadline.hour).zfill(2) + ":" + str(self.deadline.minute).zfill(2) + ":" + str(self.deadline.second).zfill(2)
        else:
            result += ", " + str(self.deadline)
        if self.location != None and self.location != "":
            result += ", " + self.location
        if self.notes != None and self.notes != "":
            result += ", " + self.notes
        return result

    def append2File(self, file, sheetName):
        wb = openpyxl.load_workbook(filename=file)
        sheet = wb[sheetName]
        allCells = sheet.calculate_dimension()
        sheet.move_range(allCells, 1, 0)
        sheet.move_range('A2:E2', -1, 0)
        sheet['A2'] = self.name
        sheet['B2'] = self.category
        sheet['C2'] = self.deadline
        sheet['D2'] = self.location
        sheet['E2'] = self.notes
        wb.save(file)
        
    @staticmethod
    def readInFromFile(rowIdx, file, sheetName):
        wb = openpyxl.load_workbook(filename=file)
        sheet = wb[sheetName]
        iterRows = sheet.iter_rows()
        row = sheet[rowIdx]
        cellsInRow = []
        for i in range(5):
            cellsInRow.append(row[i].value)
        return Assignment(cellsInRow[0], cellsInRow[1], cellsInRow[2], cellsInRow[3], cellsInRow[4])
        
    @staticmethod
    def readInAll(file, sheetName):
        rowIdx = 2
        assignment = Assignment.readInFromFile(rowIdx, file, sheetName)
        result = []
        while not assignment.none:
            result.append(assignment)
            rowIdx += 1
            assignment = Assignment.readInFromFile(rowIdx, file, sheetName)
        return result