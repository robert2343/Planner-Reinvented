import helper
import datetime
import openpyxl

class Project:
    def __init__(self, _name, _deadline, _notes):
        self.name = str(_name)
        self.deadline = _deadline
        self.notes = str(_notes)
        self.none = _name == None

    def __str__(self):
        result = self.name
        if self.deadline != None:
            if datetime.datetime.now() > self.deadline - datetime.timedelta(days = 7):
                result += ", " + helper.wdts(self.deadline.weekday()) + " " + str(self.deadline.hour).zfill(2) + ":" + str(self.deadline.minute).zfill(2) + ":" + str(self.deadline.second).zfill(2)
            else:
                result += ", " + str(self.deadline)
        if self.notes != None and self.notes != "" and self.notes != 'None':
            result += ", " + self.notes
        return result

    def append2File(self, file, sheetName):
        wb = openpyxl.load_workbook(filename=file)
        sheet = wb[sheetName]
        allCells = sheet.calculate_dimension()
        sheet.move_range(allCells, 1, 0)
        sheet.move_range('A2:C2', -1, 0)
        sheet['A2'] = self.name
        sheet['B2'] = self.deadline
        sheet['C2'] = self.notes
        wb.save(file)

    @staticmethod
    def readInFromFile(rowIdx, file, sheetName):
        wb = openpyxl.load_workbook(filename=file)
        sheet = wb[sheetName]
        iterRows = sheet.iter_rows()
        row = sheet[rowIdx]
        cellsInRow = []
        for i in range(3):
            cellsInRow.append(row[i].value)
        return Project(cellsInRow[0], cellsInRow[1], cellsInRow[2])

    @staticmethod
    def readInAll(file, sheetName):
        rowIdx = 2
        project = Project.readInFromFile(rowIdx, file, sheetName)
        result = []
        while not project.none:
            result.append(project)
            rowIdx += 1
            project = Project.readInFromFile(rowIdx, file, sheetName)
        return result
