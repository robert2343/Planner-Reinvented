import openpyxl
import datetime

class DailyTask:
    def __init__(self, _name, _pattern, _notes, dataFile, dataSheetName):
        self.name = str(_name)
        self.pattern = str(_pattern)
        self.notes = str(_notes)
        self.datesArr = []
        self.boolArr = []
        if dataFile != None:
            self.readInData(dataFile, dataSheetName)
        self.none = _name == None

    def __str__(self):
        result = self.name + ", " + self.pattern
        if self.notes != None and self.notes != "":
            result += ", " + self.notes
        result += self.calculateData()
        return result

    def calculateData(self):
        result = "\n\tStart Date: " + str(self.datesArr[0])
        if len(self.boolArr) > 9:
            result += "\n\tLast 10 times: " + str(self.boolArr[-10:])
        elif len(self.boolArr) > 0 and len(self.boolArr) <= 9:
            result += "\n\tLast " + str(len(self.boolArr)) + " times: " + str(self.boolArr[-1*len(self.boolArr):])
        numSuccess = 0
        numFail = 0
        for i in self.boolArr:
            if i == True:
                numSuccess += 1
            elif i == False:
                numFail += 1
        if numSuccess + numFail == 0:
            result += "\n\tThis is not yet begun"
        else:
            result += "\n\tPercent success: " +  str(numSuccess / (numSuccess + numFail) * 100) +  " %"
        return result


    def readInData(self, file, sheetName):
        wb = openpyxl.load_workbook(filename=file)
        sheet = wb[sheetName]
        iterRows = sheet.iter_rows()
        listRows = list(iterRows)
        dataRow = None
        dataRowIdx = None
        for i in range(1, len(listRows)):
            if listRows[i][0].value == self.name:
                dataRow = listRows[i]
                dataRowIdx = i
        try:
            dateRow = listRows[dataRowIdx - 1]
            for i in range(1, len(dataRow)):
                self.boolArr.append(dataRow[i].value)
                self.datesArr.append(dateRow[i].value)
            boolRng = range(len(self.boolArr))
            dateRng = range(len(self.datesArr))
            for i in boolRng:
                try:
                    self.boolArr.remove(None)
                except:
                    pass
            for i in dateRng:
                try:
                    self.datesArr.remove(None)
                except:
                    pass
        except:
            pass

    def calculateNextDate(self, startingPoint):
        parts = self.pattern.split()
        if self.pattern == "daily":
            return startingPoint
        elif parts[0] == "weekly":
            result = startingPoint
            while(not str(result.weekday()) in parts):
                result += datetime.timedelta(days=1)
            return result
        elif parts[0] == "monthly":
            result = startingPoint
            while(not str(result.day) in parts):
                result += datetime.timedelta(days=1)
            return result
        elif parts[0] == "quarterly":
            beginOfQtr = startingPoint
            while beginOfQtr.replace(hour=0, minute=0, second=0, microsecond=0) != datetime.datetime(year=beginOfQtr.year, month=1, day=1) and beginOfQtr.replace(hour=0, minute=0, second=0, microsecond=0) != datetime.datetime(year=beginOfQtr.year, month=4, day=1) and beginOfQtr.replace(hour=0, minute=0, second=0, microsecond=0) != datetime.datetime(year=beginOfQtr.year, month=7, day=1) and beginOfQtr.replace(hour=0, minute=0, second=0, microsecond=0) != datetime.datetime(year=beginOfQtr.year, month=10, day=1):
                beginOfQtr -= datetime.timedelta(days=1)
            result = beginOfQtr
            count = 1
            while startingPoint > result or not str(count) in parts:
                count += 1
                result += datetime.timedelta(days=1)
                if result.replace(hour=0, minute=0, second=0, microsecond=0) == datetime.datetime(year=result.year, month=1, day=1) or result.replace(hour=0, minute=0, second=0, microsecond=0) == datetime.datetime(year=result.year, month=4, day=1) or result.replace(hour=0, minute=0, second=0, microsecond=0) == datetime.datetime(year=result.year, month=7, day=1) or result.replace(hour=0, minute=0, second=0, microsecond=0) == datetime.datetime(year=result.year, month=10, day=1):
                    count = 1
            return result
        elif parts[0] == "yearly":
            beginOfYear = startingPoint.replace(month=1, day=1)
            result = beginOfYear
            count = 1
            lastYear = result.year
            while startingPoint > result or not str(count) in parts:
                count += 1
                result += datetime.timedelta(days=1)
                if lastYear != result.year:
                    count = 1
                    lastYear = result.year
            return result
        elif parts[0] == "every":
            return startingPoint

    def append2File(self, file, sheetName, dataSheetName):
        wb = openpyxl.load_workbook(filename=file)
        sheet = wb[sheetName]
        allCells = sheet.calculate_dimension()
        sheet.move_range(allCells, 1, 0)
        sheet.move_range('A2:C2', -1, 0)
        sheet['A2'] = self.name
        sheet['B2'] = self.pattern
        sheet['C2'] = self.notes
        sheet = wb[dataSheetName]
        allCells = sheet.calculate_dimension()
        sheet.move_range(allCells, 2, 0)
        sheet.move_range('A3:B3', -2, 0)
        sheet['A3'] = self.name
        sheet['B2'] = self.calculateNextDate(datetime.datetime.now())
        wb.save(file)

    def runToday(self):
        pass

    @staticmethod
    def readInFromFile(rowIdx, file, sheetName, dataSheetName):
        wb = openpyxl.load_workbook(filename=file)
        sheet = wb[sheetName]
        row = sheet[rowIdx]
        cellsInRow = []
        for i in range(3):
            cellsInRow.append(row[i].value)
        return DailyTask(cellsInRow[0], cellsInRow[1], cellsInRow[2], file, dataSheetName)

    @staticmethod
    def readInAll(file, sheetName, dataSheetName):
        rowIdx = 2
        assignment = DailyTask.readInFromFile(rowIdx, file, sheetName, dataSheetName)
        result = []
        while not assignment.none:
            result.append(assignment)
            rowIdx += 1
            assignment = DailyTask.readInFromFile(rowIdx, file, sheetName, dataSheetName)
        return result
